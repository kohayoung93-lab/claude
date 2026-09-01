#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
계정별원장 자동화 스크립트 (공통양식 기반)
--------------------------------
[재무상태표], [계정] 시트가 있는 템플릿 워크북과, ERP(이카운트)에서 받은
raw 원장 파일들을 이용해 계정코드별로 "계정별원장" 시트를 하나씩 만들어 주는
도구입니다.

폴더 구조 (이 파일과 같은 위치):
  master/     - [재무상태표], [계정] 시트가 포함된 템플릿 파일 1개
  input_raw/  - ERP에서 받은 raw 원장 xlsx 파일들 (여러 개 가능)
  output/     - 결과가 저장되는 곳 (템플릿 원본은 수정되지 않음)

사용법:
  python3 build_ledger.py
(run_mac.command / run_windows.bat 더블클릭으로도 실행 가능)

동작 방식:
 - 계정별 원장 시트마다 G1=계정코드, H1=계정과목, I1=Raw 마지막잔액, J1=재무제표 일치(TRUE/FALSE)를
   "고정 위치" 수식으로 둔다. I1은 그 시트의 '합계' 행을 MATCH로 찾아오므로, 데이터 행 수가
   달라져도(Raw를 통째로 갈아끼워도) 항상 정확한 값을 찾는다. -> 무거운 배열식/전체범위 SUMPRODUCT 대신
   가벼운 INDEX/MATCH 한 줄만 사용.
 - [계정] 시트가 맨 앞 탭으로 온다.
 - [계정] 시트 C열(Raw유무)/D열(재무제표 일치)도 각 원장 시트의 I1/J1을 그대로 참조하는
   가벼운 단일 셀 참조 수식으로 바뀐다 (매크로/INDIRECT 없음 -> 안 무거움).
 - Raw가 아직 없는 계정도 "빈 템플릿 시트"를 미리 만들어 둔다. 앞으로는 Raw 파일을 다시
   올릴 필요 없이, 해당 계정 시트의 3행부터 원장 데이터를 붙여넣기만 하면 I1/J1/[계정]시트가
   자동으로 갱신된다.

주의: 계정코드가 같은데 [계정] 시트에 두 번 등장하는 경우(예: 유동/비유동에 같은 계정명이 각각
있는 케이스)는 시트가 1개만 만들어지고, 그 시트의 J1은 "당기 금액이 있는 재무상태표 행" 기준
하나로만 판정된다. 나머지 한 쪽 행은 사람이 직접 확인해야 한다. (아래 콘솔 경고로 표시)
"""
import datetime
import glob
import os
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "input_raw")
MASTER_DIR = os.path.join(BASE_DIR, "master")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

TITLE_RE = re.compile(r"계정별원장\s*/\s*(\d+)\((.+)\)\s*$")
NUM_FMT = "#,##0"
HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
META_FILL = PatternFill("solid", fgColor="FFF2CC")
RAW_HEADERS = ["일자-No.", "계정코드", "계정명", "거래처코드", "거래처명", "적요",
               "차변금액", "대변금액", "잔액", "상대계정코드", "상대계정명",
               "부서명", "최종수정자", "최종수정일자"]
COL_WIDTHS = [18, 8, 14, 14, 16, 26, 18, 18, 18, 10, 14, 12, 10, 20]


def find_files(dir_path, label, required=True):
    files = [f for f in glob.glob(os.path.join(dir_path, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if required and len(files) == 0:
        raise RuntimeError(f"{label} 폴더({dir_path})에 xlsx 파일이 없습니다.")
    return files


def sanitize_sheet_name(name, existing):
    bad = set('[]:*?/\\')
    clean = "".join(c for c in str(name) if c not in bad).strip()
    clean = clean[:31] if clean else "SHEET"
    base = clean
    i = 2
    while clean in existing:
        suffix = f"_{i}"
        clean = (base[: 31 - len(suffix)]) + suffix
        i += 1
    existing.add(clean)
    return clean


def load_account_list(ws_acc):
    accounts = []
    header_row = None
    for r in range(1, ws_acc.max_row + 1):
        if ws_acc.cell(r, 1).value == "계정코드" and ws_acc.cell(r, 2).value == "재무제표표시명":
            header_row = r
            break
    if header_row is None:
        raise ValueError("[계정] 시트에서 헤더행('계정코드','재무제표표시명')을 찾지 못했습니다.")
    for r in range(header_row + 1, ws_acc.max_row + 1):
        code = ws_acc.cell(r, 1).value
        name = ws_acc.cell(r, 2).value
        if code is None and name is None:
            continue
        accounts.append((r, code, name))
    return header_row, accounts


def normalize_name(name):
    """공백류 문자(스페이스/탭/줄바꿈/전각공백 등) 차이를 무시하기 위해 전부 제거한 이름."""
    if not isinstance(name, str):
        return name
    return re.sub(r"\s+", "", name)


def fill_bs_total_column(ws_bs_formula):
    """재무상태표 F열(당기 금액) = B열 + C열 수식을 A열에 계정명이 있는 모든 행에 채운다."""
    for r in range(1, ws_bs_formula.max_row + 1):
        name = ws_bs_formula.cell(r, 1).value
        if not name or not isinstance(name, str):
            continue
        ws_bs_formula.cell(r, 6, f"=B{r}+C{r}")


def build_bs_map(ws_bs_formula, ws_bs_value):
    """재무상태표 A열(계정명) -> row. 동일 이름이 여러 행이면 당기(B+C)값이
    존재/0이 아닌 행을 우선 채택(그래도 안되면 첫 행)."""
    name_rows = {}
    for r in range(1, ws_bs_formula.max_row + 1):
        name = ws_bs_formula.cell(r, 1).value
        if not name or not isinstance(name, str):
            continue
        name_rows.setdefault(normalize_name(name), []).append(r)

    def bc_sum(r):
        b = ws_bs_value.cell(r, 2).value
        c = ws_bs_value.cell(r, 3).value
        b = b if isinstance(b, (int, float)) else 0
        c = c if isinstance(c, (int, float)) else 0
        return b + c

    resolved, ambiguous = {}, {}
    for name, rows in name_rows.items():
        if len(rows) == 1:
            resolved[name] = rows[0]
        else:
            nonzero = [r for r in rows if bc_sum(r) != 0]
            resolved[name] = nonzero[0] if len(nonzero) == 1 else rows[0]
            ambiguous[name] = rows
    return resolved, ambiguous


def parse_raw_blocks(path):
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    max_col, max_row = ws.max_column, ws.max_row

    title_rows = []
    for r in range(1, max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and "계정별원장" in v and "회사명" in v:
            m = TITLE_RE.search(v)
            if m:
                title_rows.append((r, m.group(1), m.group(2)))

    blocks = {}
    for idx, (start_r, code, raw_name) in enumerate(title_rows):
        end_r = title_rows[idx + 1][0] - 1 if idx + 1 < len(title_rows) else max_row
        total_r = None
        for r in range(start_r, end_r + 1):
            if ws.cell(r, 1).value == "합계":
                total_r = r
                break
        if total_r is None:
            total_r = end_r
        rows = [tuple(ws.cell(r, c).value for c in range(1, max_col + 1))
                for r in range(start_r, total_r + 1)]
        blocks[code] = {"name": raw_name, "rows": rows, "src": path}
    return blocks


def style_ledger_sheet(ws, header_row_idx=2, data_last_row=None):
    bold = Font(bold=True)
    for c_i in range(1, len(RAW_HEADERS) + 1):
        cell = ws.cell(header_row_idx, c_i)
        cell.font = bold
        cell.fill = HEADER_FILL
    # 컬럼 단위로 숫자서식을 지정 (셀 단위 반복보다 훨씬 가볍고, 나중에
    # 붙여넣는 새 행에도 자동 적용됨). 실제 데이터가 있는 시트는 아래에서
    # 셀 단위로도 한 번 더 지정해서(콤마가 확실히 보이도록) 덮어쓴다.
    for c_i in (7, 8, 9):
        ws.column_dimensions[get_column_letter(c_i)].number_format = NUM_FMT
    for c_i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c_i)].width = w
    ws.freeze_panes = "A3"

    if data_last_row:
        # 차변/대변/잔액 열: 셀 단위 콤마서식 적용 + 실제 숫자 길이에 맞춰 열 너비 자동조정
        max_len = {7: len(RAW_HEADERS[6]), 8: len(RAW_HEADERS[7]), 9: len(RAW_HEADERS[8])}
        for r in range(header_row_idx + 1, data_last_row + 1):
            for c_i in (7, 8, 9):
                cell = ws.cell(r, c_i)
                cell.number_format = NUM_FMT
                if isinstance(cell.value, (int, float)):
                    s = format(cell.value, ",")
                    if len(s) > max_len[c_i]:
                        max_len[c_i] = len(s)
        for c_i, ln in max_len.items():
            ws.column_dimensions[get_column_letter(c_i)].width = max(ln + 2, COL_WIDTHS[c_i - 1])


def write_meta_row1(ws, code, name, bs_row):
    """G1/H1/I1/J1 고정 메타 셀. I1/J1은 데이터 행 수와 무관하게 항상 맞는 값을
    찾도록 INDEX/MATCH(가벼움, 비휘발성)만 사용한다.
    주의: '합계' 행의 잔액(I열)이 정확히 0이면 원본 Raw 자체가 그 칸을 공란으로
    내보내는 경우가 있어, "값이 있는지"(ISNUMBER) 대신 "합계 행 자체가 있는지"
    (COUNTIF)로 Raw 유무를 판정한다 - 잔액 0인 계정도 Raw있음으로 정확히 잡힌다."""
    ws.cell(1, 7, code)   # G1 계정코드
    ws.cell(1, 8, name)   # H1 계정과목
    ws.cell(1, 9, '=IF(COUNTIF(A:A,"합계")=0,"",INDEX(I:I,MATCH("합계",A:A,0)))')  # I1
    if bs_row:
        # N()으로 숫자화: 합계행의 잔액이 정확히 0이라 원본 셀이 공란인 경우에도
        # "빈 셀=""와 같다"는 Excel 특성 때문에 I1=""이 True로 오판되는 것을 방지.
        ws.cell(1, 10, f'=IF(COUNTIF(A:A,"합계")=0,"",IF(N(I1)=N(재무상태표!F{bs_row}),TRUE,FALSE))')  # J1
    else:
        ws.cell(1, 10, "재무상태표 미매칭")
    ws.cell(1, 12, '=HYPERLINK("#재무상태표!A1","home")')  # L1: 재무상태표로 바로가기
    ws.cell(1, 12).font = Font(bold=True, color="0563C1", underline="single")
    for c_i in (7, 8, 9, 10):
        ws.cell(1, c_i).fill = META_FILL
        ws.cell(1, c_i).font = Font(bold=True)
    ws.cell(1, 9).number_format = NUM_FMT


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    master_files = find_files(MASTER_DIR, "master")
    if len(master_files) > 1:
        raise RuntimeError(f"master 폴더에는 파일이 1개만 있어야 합니다. 현재: {[os.path.basename(f) for f in master_files]}")
    template_path = master_files[0]
    print(f"[템플릿] {os.path.basename(template_path)}")

    raw_files = find_files(RAW_DIR, "raw", required=False)
    print(f"[raw 파일 {len(raw_files)}개]")
    for f in raw_files:
        print(f"  - {os.path.basename(f)}")
    print()

    wb = openpyxl.load_workbook(template_path, data_only=False)
    wb_val = openpyxl.load_workbook(template_path, data_only=True)

    if "재무상태표" not in wb.sheetnames or "계정" not in wb.sheetnames:
        print("템플릿에 [재무상태표], [계정] 시트가 모두 있어야 합니다.", file=sys.stderr)
        sys.exit(1)

    ws_acc = wb["계정"]
    ws_bs = wb["재무상태표"]
    ws_bs_val = wb_val["재무상태표"]

    fill_bs_total_column(ws_bs)

    header_row, accounts = load_account_list(ws_acc)
    bs_map, bs_ambiguous = build_bs_map(ws_bs, ws_bs_val)

    all_blocks = {}
    dup_code_warnings = []
    for path in raw_files:
        for code, b in parse_raw_blocks(path).items():
            if code in all_blocks:
                dup_code_warnings.append((code, all_blocks[code]["src"], b["src"]))
            all_blocks[code] = b

    known_codes = {str(code) for _, code, _ in accounts if code is not None}
    unmatched_raw_codes = [c for c in all_blocks if c not in known_codes]

    bold = Font(bold=True)
    ws_acc.cell(header_row, 3, "Raw유무")
    ws_acc.cell(header_row, 4, "재무제표 일치")
    for c in (3, 4):
        ws_acc.cell(header_row, c).font = bold
        ws_acc.cell(header_row, c).fill = HEADER_FILL
    ws_acc.column_dimensions["C"].width = 12
    ws_acc.column_dimensions["D"].width = 16

    # 코드 하나당 시트 1개 (계정 시트에 같은 코드가 중복 등장해도 시트는 재사용)
    existing_sheet_names = set(wb.sheetnames) - {"재무상태표", "계정"}
    code_to_sheet = {}
    dup_code_rows = {}  # code -> [row_no,...] 2개 이상이면 경고용

    for row_no, code, name in accounts:
        code_str = str(code) if code is not None else None
        dup_code_rows.setdefault(code_str, []).append(row_no)

    for row_no, code, name in accounts:
        code_str = str(code) if code is not None else None
        bs_row = bs_map.get(normalize_name(name))

        if code_str in code_to_sheet:
            sheet_name = code_to_sheet[code_str]
        else:
            sheet_name = sanitize_sheet_name(name, existing_sheet_names)
            code_to_sheet[code_str] = sheet_name
            ws_new = wb.create_sheet(sheet_name)

            block = all_blocks.get(code_str)
            if block:
                for r_i, row_vals in enumerate(block["rows"], start=1):
                    for c_i, val in enumerate(row_vals, start=1):
                        ws_new.cell(r_i, c_i, val)
                style_ledger_sheet(ws_new, header_row_idx=2, data_last_row=len(block["rows"]))
            else:
                ws_new.cell(1, 1, "※ 3행부터 계정별원장 Raw 데이터를 그대로 붙여넣어 주세요 (2행 머리글 포함)")
                for c_i, h in enumerate(RAW_HEADERS, start=1):
                    ws_new.cell(2, c_i, h)
                style_ledger_sheet(ws_new)

            write_meta_row1(ws_new, code, name, bs_row)

        ws_acc.cell(row_no, 3, f"=IF(COUNTIF('{sheet_name}'!A:A,\"합계\")=0,\"Raw없음\",\"Raw있음\")")
        ws_acc.cell(row_no, 4, f"=IFERROR('{sheet_name}'!J1,\"\")")

    # ---- Raw없음 / FALSE 강조 (조건부서식: 수식 결과라도 값이 바뀔 때마다 자동 재적용됨) ----
    red_font = Font(color="FF0000")
    first_data_row = accounts[0][0]
    last_data_row = accounts[-1][0]
    c_range = f"C{first_data_row}:C{last_data_row}"
    d_range = f"D{first_data_row}:D{last_data_row}"
    ws_acc.conditional_formatting.add(
        c_range, CellIsRule(operator="equal", formula=['"Raw없음"'], font=red_font)
    )
    ws_acc.conditional_formatting.add(
        d_range, CellIsRule(operator="equal", formula=["FALSE"], font=red_font)
    )

    # ---- 시트 순서 재배열: 계정 -> 재무상태표 -> 계정별 시트들(계정시트 등장 순) ----
    ordered_names = ["계정", "재무상태표"] + list(dict.fromkeys(code_to_sheet.values()))
    remaining = [n for n in wb.sheetnames if n not in ordered_names]
    final_order = ordered_names + remaining
    wb._sheets = [wb[n] for n in final_order]
    wb.active = 0

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    template_stem = os.path.splitext(os.path.basename(template_path))[0]
    output_path = os.path.join(OUTPUT_DIR, f"{template_stem}_계정별원장_{timestamp}.xlsx")
    wb.save(output_path)

    n_total = len(dup_code_rows)
    n_with_raw = len(set(code_to_sheet[c] for c in code_to_sheet if c in all_blocks))
    print(f"완료: {output_path}")
    print(f" - 총 계정(코드) 수: {n_total}, 시트 생성: {len(code_to_sheet)}")
    print(f" - Raw 있음: {len(all_blocks)}, Raw 없음(빈 템플릿): {len(code_to_sheet) - len(all_blocks)}")
    if dup_code_warnings:
        print(" [경고] 동일 계정코드가 여러 Raw 파일에 중복 존재(나중 파일로 덮어씀):")
        for code, f1, f2 in dup_code_warnings:
            print(f"   - 코드 {code}: {f1} / {f2}")
    if unmatched_raw_codes:
        print(" [경고] [계정] 시트에 없는 계정코드가 Raw에 존재:", unmatched_raw_codes)
    dup_codes_in_acc = {c: rs for c, rs in dup_code_rows.items() if len(rs) > 1}
    if dup_codes_in_acc:
        print(" [주의] [계정] 시트에 같은 계정코드가 2번 이상 등장 (시트/판정은 1개만 생성됨, 수동 확인 필요):")
        for c, rs in dup_codes_in_acc.items():
            print(f"   - 코드 {c}: 행 {rs}")
    if bs_ambiguous:
        print(" [참고] 재무상태표에 동일 계정명이 여러 행 존재 (당기값 있는 행을 자동 채택):")
        for name, rows in bs_ambiguous.items():
            print(f"   - {name}: 행 {rows}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[오류] {e}")
        sys.exit(1)
