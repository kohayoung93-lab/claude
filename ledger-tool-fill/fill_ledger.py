#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
raw 원장 파일들을, "데이터는 없고 시트/제목/헤더/피벗만 남아있는" master 워크북의
각 회사 시트에 A~V열만 채워 넣는 도구.

ledger-tool(누적형) / ledger-tool-new(신규 생성)와의 차이:
  - 이 도구는 이미 존재하는 큰 워크북(master) 안의 "비어있는" 여러 시트를
    한 번에 채운다. 시트가 비어있으면(이미 실거래가 없으면) 이월잔액을
    포함하고, 이미 데이터가 있는 시트라면(재실행 등) 이월잔액은 제외한다
    (ledger-tool과 동일한 이유 - 이미 반영되어 있으므로).
  - A~V열(일자-No ~ 상대계정명)만 raw에서 그대로 채우고, W열(상대거래처코드)
    ~AC열(비고)은 비워둔다. 단 Y(잔액2)/Z(구분1)/AA(구분2)는 ledger-tool과
    동일하게 살아있는 수식(SUMIFS, INDEX/MATCH)으로 채운다.
  - master에 있던 피벗테이블은 전부 제거한다 (관련 캐시/관계/헤더 잔여물 포함).
  - 1행(제목+합계 수식)과 2행(헤더)은 절대 건드리지 않는다.

폴더 구조 (이 파일과 같은 위치):
  input_raw/  - raw 원장 xlsx 파일들. 파일명 = 채울 시트 이름 (예: HD.xlsx -> "HD" 시트)
  master/     - 채워 넣을 대상 워크북 1개 (여러 시트 + 계정과목표 시트 포함)
  output/     - 결과가 저장되는 곳 (master 원본은 건드리지 않음)

사용법:
  python3 fill_ledger.py
(run_mac.command / run_windows.bat 더블클릭으로도 실행 가능)
"""

import datetime
import glob
import os
import re
import sys
import zipfile

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "input_raw")
MASTER_DIR = os.path.join(BASE_DIR, "master")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

MAPPING_SHEET_NAME = "99.이카운트 계정과목표_HF"

RAW_COLS = [
    "일자-No", "계정코드", "계정명", "거래처코드", "거래처명", "적요",
    "차변금액", "대변금액", "잔액", "부서명", "최초작성일자", "최초작성자",
    "최종수정일자", "최종수정자", "채권채무번호", "만기일자", "은행", "계좌번호",
    "예금주명", "상대계정코드", "상대계정명", "상대거래처코드", "상대거래처명",
]
DATE_NO_RE = re.compile(r"^\d{4}/\d{2}/\d{2}\s*-")
PERIOD_RE = re.compile(r"(\d{4}/\d{2}/\d{2})\s*~\s*(\d{4}/\d{2}/\d{2})")
EXCEL_EPOCH = datetime.date(1899, 12, 30)

COLS_AV = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
           "O", "P", "Q", "R", "S", "T", "U", "V"]


def esc(s):
    return (
        str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
    )


def norm_code(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        return str(int(float(s)))
    except ValueError:
        return s


def excel_date_serial(date_str):
    if not date_str:
        return None
    try:
        y, m, d = (int(x) for x in date_str.split("/"))
        return (datetime.date(y, m, d) - EXCEL_EPOCH).days
    except (ValueError, TypeError):
        return None


def normalize_date_str(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y/%m/%d")
    return str(v).strip()


# ---------------------------------------------------------------------------
# 1. 매핑표 로딩
# ---------------------------------------------------------------------------

def load_mapping(master_path):
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    if MAPPING_SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"master 파일에서 '{MAPPING_SHEET_NAME}' 시트를 찾지 못했습니다.")
    ws = wb[MAPPING_SHEET_NAME]
    mapping = {}
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        code = norm_code(row[0])
        if code is None:
            continue
        mapping[code] = {"구분1": row[4], "구분2": row[3]}
    wb.close()
    return mapping


# ---------------------------------------------------------------------------
# 2. raw 파일 파싱
# ---------------------------------------------------------------------------

def parse_raw_file(raw_path, include_ol):
    """include_ol=True면 이월잔액을 기초잔액 거래로 포함, False면 제외
    (대상 시트에 이미 데이터가 있어 이미 반영되어 있다고 볼 때)."""
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    ws = wb.worksheets[0]

    transactions = []
    blocks_report = []
    header_idx = None
    block = None
    title_code, title_name = None, None
    period = None
    title_re = re.compile(r"(\d+)\(([^)]+)\)\s*$")

    def flush_block():
        if block is not None:
            blocks_report.append(block)

    for row in ws.iter_rows(values_only=True):
        first_cell = row[0]
        if isinstance(first_cell, str) and first_cell.startswith("회사명"):
            if period is None:
                pm = PERIOD_RE.search(first_cell)
                if pm:
                    period = (pm.group(1), pm.group(2))
            m = title_re.search(first_cell)
            if m:
                title_code, title_name = m.group(1), m.group(2)
            continue

        if isinstance(first_cell, str) and first_cell.strip() == "일자-No":
            flush_block()
            header_idx = {name: i for i, name in enumerate(row) if name}
            block = {
                "계정코드": title_code, "계정명": title_name,
                "kept_debit": 0, "kept_credit": 0, "ol_debit": 0, "ol_credit": 0,
                "stated_debit": None, "stated_credit": None, "has_summary_row": False,
            }
            continue

        if header_idx is None:
            continue

        no_val = first_cell
        row_jeokyo = row[header_idx["적요"]] if "적요" in header_idx else None
        is_ol = isinstance(row_jeokyo, str) and row_jeokyo.strip() == "이월잔액"

        if is_ol:
            rec = {name: row[header_idx[name]] if name in header_idx else None for name in RAW_COLS}
            block["ol_debit"] += rec["차변금액"] or 0
            block["ol_credit"] += rec["대변금액"] or 0
            if include_ol:
                if period:
                    rec["일자-No"] = f"{period[0]} -000"
                rec["계정코드"] = rec["계정코드"] or title_code
                rec["계정명"] = rec["계정명"] or title_name
                transactions.append(rec)
                block["kept_debit"] += rec["차변금액"] or 0
                block["kept_credit"] += rec["대변금액"] or 0
        elif isinstance(no_val, str) and DATE_NO_RE.match(no_val):
            rec = {name: row[header_idx[name]] if name in header_idx else None for name in RAW_COLS}
            transactions.append(rec)
            block["kept_debit"] += rec["차변금액"] or 0
            block["kept_credit"] += rec["대변금액"] or 0
        elif isinstance(no_val, str) and no_val.strip() == "합계":
            di = header_idx.get("차변금액")
            ci = header_idx.get("대변금액")
            block["stated_debit"] = row[di] if di is not None else None
            block["stated_credit"] = row[ci] if ci is not None else None
            block["has_summary_row"] = True

    flush_block()
    wb.close()
    return transactions, blocks_report, period


def validate_blocks(blocks_report, include_ol):
    results = []
    for b in blocks_report:
        if include_ol:
            if not b["has_summary_row"]:
                expected_d, expected_c = b["ol_debit"], b["ol_credit"]
            else:
                expected_d, expected_c = b["stated_debit"] or 0, b["stated_credit"] or 0
        else:
            if not b["has_summary_row"]:
                expected_d, expected_c = 0, 0
            else:
                expected_d = (b["stated_debit"] or 0) - b["ol_debit"]
                expected_c = (b["stated_credit"] or 0) - b["ol_credit"]
        ok = (expected_d == b["kept_debit"]) and (expected_c == b["kept_credit"])
        results.append({
            "계정코드": b["계정코드"], "계정명": b["계정명"], "ok": ok,
            "kept_debit": b["kept_debit"], "kept_credit": b["kept_credit"],
            "expected_debit": expected_d, "expected_credit": expected_c,
        })
    return results


# ---------------------------------------------------------------------------
# 3. master의 대상 시트 현재 상태 읽기
# ---------------------------------------------------------------------------

def read_target_sheet_state(master_path, sheet_name):
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"master 파일에 '{sheet_name}' 시트가 없습니다. 사용 가능한 시트: {wb.sheetnames}")
    ws = wb[sheet_name]

    last_data_row = 2  # 헤더 행. 실거래가 없으면 그대로 유지된다.
    running = {}
    last_date_str = None

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        a = row[0]
        if isinstance(a, str) and DATE_NO_RE.match(a):
            last_data_row = i
            last_date_str = normalize_date_str(row[1])
            acct_name, debit, credit = row[3], row[7], row[8]
            cur = running.get(acct_name, [0, 0])
            cur[0] += (debit or 0)
            cur[1] += (credit or 0)
            running[acct_name] = cur

    wb.close()
    return {"last_data_row": last_data_row, "running": running, "last_date_str": last_date_str}


# ---------------------------------------------------------------------------
# 4. workbook 안의 sheet xml 경로 찾기
# ---------------------------------------------------------------------------

def find_sheet_xml_path(zip_entries, sheet_name):
    wb_xml = zip_entries["xl/workbook.xml"].decode("utf-8")
    m = re.search(rf'<sheet name="{re.escape(sheet_name)}"[^>]*r:id="(rId\d+)"', wb_xml)
    if not m:
        raise RuntimeError(f"workbook.xml에서 '{sheet_name}' 시트를 찾지 못했습니다.")
    rels_xml = zip_entries["xl/_rels/workbook.xml.rels"].decode("utf-8")
    m2 = re.search(rf'<Relationship Id="{m.group(1)}"[^>]*Target="([^"]+)"', rels_xml)
    return "xl/" + m2.group(1).lstrip("/")


# ---------------------------------------------------------------------------
# 5. 데이터 행용 스타일 준비 (styles.xml에 없으면 새로 추가, 워크북 전체에서 1회만)
# ---------------------------------------------------------------------------

def _parse_xf_list(inner_xml):
    return re.findall(r"<xf[^/]*/>|<xf[^>]*>.*?</xf>", inner_xml, re.S)


def ensure_data_styles(zip_entries):
    """데이터 행에 쓸 4가지 스타일(왼쪽정렬/가운데정렬/금액/날짜)이 styles.xml에
    이미 있으면 그 인덱스를 재사용하고, 없으면 새로 추가한다 (재실행 시 중복 추가
    방지)."""
    styles_xml = zip_entries["xl/styles.xml"].decode("utf-8")

    fm = re.search(r'<fonts count="(\d+)"([^>]*)>(.*?)</fonts>', styles_xml, re.S)
    fonts_inner = fm.group(3)
    font_list = re.findall(r"<font>.*?</font>", fonts_inner, re.S)
    marker_font = '<font><sz val="10"/><name val="Arial"/><family val="2"/></font>'
    if marker_font in font_list:
        font_id = font_list.index(marker_font)
    else:
        font_id = len(font_list)
        fonts_inner += marker_font
        styles_xml = (
            styles_xml[:fm.start()]
            + f'<fonts count="{len(font_list) + 1}"{fm.group(2)}>{fonts_inner}</fonts>'
            + styles_xml[fm.end():]
        )

    xm = re.search(r'<cellXfs count="(\d+)">(.*?)</cellXfs>', styles_xml, re.S)
    xfs_inner = xm.group(2)
    xf_list = _parse_xf_list(xfs_inner)
    BORDER_ID = 11
    wanted = [
        ("left", f'<xf numFmtId="0" fontId="{font_id}" fillId="0" borderId="{BORDER_ID}" xfId="0" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="left" vertical="center"/></xf>'),
        ("center", f'<xf numFmtId="0" fontId="{font_id}" fillId="0" borderId="{BORDER_ID}" xfId="0" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>'),
        ("money", f'<xf numFmtId="3" fontId="{font_id}" fillId="0" borderId="{BORDER_ID}" xfId="0" applyNumberFormat="1" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="right" vertical="center"/></xf>'),
        ("date", f'<xf numFmtId="14" fontId="{font_id}" fillId="0" borderId="{BORDER_ID}" xfId="0" applyNumberFormat="1" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center"/></xf>'),
    ]
    style_ids = {}
    new_xfs = []
    next_id = len(xf_list)
    for name, xf_text in wanted:
        if xf_text in xf_list:
            style_ids[name] = xf_list.index(xf_text)
        else:
            style_ids[name] = next_id
            new_xfs.append(xf_text)
            xf_list.append(xf_text)
            next_id += 1

    if new_xfs:
        new_inner = xfs_inner + "".join(new_xfs)
        styles_xml = (
            styles_xml[:xm.start()]
            + f'<cellXfs count="{next_id}">{new_inner}</cellXfs>'
            + styles_xml[xm.end():]
        )

    zip_entries["xl/styles.xml"] = styles_xml.encode("utf-8")
    return style_ids


STYLE_COL_KIND = {
    "A": "left", "B": "date", "C": "center", "D": "left", "E": "center", "F": "left", "G": "left",
    "H": "money", "I": "money", "J": "money", "K": "left", "L": "left", "M": "left", "N": "left",
    "O": "left", "P": "left", "Q": "left", "R": "left", "S": "left", "T": "left", "U": "center", "V": "left",
}


# ---------------------------------------------------------------------------
# 6. 셀/행 XML 생성
# ---------------------------------------------------------------------------

def build_cell(col, row_num, style, kind, value):
    ref = f"{col}{row_num}"
    if value is None or value == "":
        return f'<c r="{ref}" s="{style}"/>'
    if kind == "text":
        return f'<c r="{ref}" s="{style}" t="inlineStr"><is><t xml:space="preserve">{esc(value)}</t></is></c>'
    return f'<c r="{ref}" s="{style}"><v>{value}</v></c>'


def build_cell_formula_num(col, row_num, style, formula, cached):
    return f'<c r="{col}{row_num}" s="{style}"><f>{esc(formula)}</f><v>{cached}</v></c>'


def build_cell_formula_str(col, row_num, style, formula, cached):
    return f'<c r="{col}{row_num}" s="{style}" t="str"><f>{esc(formula)}</f><v>{esc(cached)}</v></c>'


def build_new_rows(new_row_start, records, style_ids, mapping, running, sheet_name):
    xml_parts = []
    warnings = []
    for offset, rec in enumerate(records):
        r = new_row_start + offset
        raw_no = rec["일자-No"]
        date_str = raw_no.split()[0] if raw_no else None
        acct_code_norm = norm_code(rec["계정코드"])
        acct_code_num = int(acct_code_norm) if acct_code_norm is not None else None
        acct_name = rec["계정명"]
        debit = rec["차변금액"] or 0
        credit = rec["대변금액"] or 0

        cur = running.get(acct_name, [0, 0])
        cur[0] += debit
        cur[1] += credit
        running[acct_name] = cur
        y_value = abs(cur[0] - cur[1])
        y_formula = f"ABS(SUMIFS(H$3:H{r}, D$3:D{r}, D{r}) - SUMIFS(I$3:I{r}, D$3:D{r}, D{r}))"

        map_entry = mapping.get(acct_code_norm)
        if map_entry is None:
            warnings.append(f"[{sheet_name}] 계정코드 {rec['계정코드']} 이(가) 매핑표에 없습니다 (행 {r}).")
            gubun1, gubun2 = "", ""
        else:
            gubun1 = map_entry["구분1"] if map_entry["구분1"] is not None else ""
            gubun2 = map_entry["구분2"] if map_entry["구분2"] is not None else ""
        z_formula = f"INDEX('{MAPPING_SHEET_NAME}'!E:E,MATCH('{sheet_name}'!C{r},'{MAPPING_SHEET_NAME}'!A:A,0))"
        aa_formula = f"INDEX('{MAPPING_SHEET_NAME}'!D:D,MATCH('{sheet_name}'!C{r},'{MAPPING_SHEET_NAME}'!A:A,0))"

        values_by_col = {
            "A": ("text", raw_no), "B": ("num", excel_date_serial(date_str)), "C": ("num", acct_code_num),
            "D": ("text", acct_name), "E": ("text", rec["거래처코드"]), "F": ("text", rec["거래처명"]),
            "G": ("text", rec["적요"]), "H": ("num", rec["차변금액"]), "I": ("num", rec["대변금액"]),
            "J": ("num", rec["잔액"]), "K": ("text", rec["부서명"]), "L": ("text", rec["최초작성일자"]),
            "M": ("text", rec["최초작성자"]), "N": ("text", rec["최종수정일자"]), "O": ("text", rec["최종수정자"]),
            "P": ("text", rec["채권채무번호"]), "Q": ("text", rec["만기일자"]), "R": ("text", rec["은행"]),
            "S": ("text", rec["계좌번호"]), "T": ("text", rec["예금주명"]), "U": ("text", rec["상대계정코드"]),
            "V": ("text", rec["상대계정명"]),
        }
        cells = []
        for col in COLS_AV:
            kind, val = values_by_col[col]
            cells.append(build_cell(col, r, style_ids[STYLE_COL_KIND[col]], kind, val))
        cells.append(build_cell_formula_num("Y", r, style_ids["money"], y_formula, y_value))
        cells.append(build_cell_formula_str("Z", r, style_ids["center"], z_formula, gubun1))
        cells.append(build_cell_formula_str("AA", r, style_ids["center"], aa_formula, gubun2))

        xml_parts.append(f'<row r="{r}" spans="1:33">' + "".join(cells) + "</row>")

    return "".join(xml_parts), warnings


# ---------------------------------------------------------------------------
# 7. 마지막 실거래 행 뒤에 남은 내용(각주 등)을 밀어내는 로직 (ledger-tool과 동일)
# ---------------------------------------------------------------------------

def shift_trailing_block(block_text, old_start, old_end, n):
    def repl_row(m):
        return f'<row r="{int(m.group(1)) + n}"'
    block_text = re.sub(r'<row r="(\d+)"', repl_row, block_text)
    nums = "|".join(str(x) for x in range(old_start, old_end + 1))
    pattern = re.compile(r"([A-Z]{1,3})(" + nums + r")\b")

    def repl_cell(m):
        return f"{m.group(1)}{int(m.group(2)) + n}"
    return pattern.sub(repl_cell, block_text)


def shift_merge_cells(sheet_xml, old_start, old_end, n):
    nums = "|".join(str(x) for x in range(old_start, old_end + 1))
    pattern = re.compile(r"([A-Z]{1,3})(" + nums + r")\b")

    def repl(m):
        return f"{m.group(1)}{int(m.group(2)) + n}"

    def repl_merge(m):
        return '<mergeCell ref="' + pattern.sub(repl, m.group(1)) + '"/>'
    return re.sub(r'<mergeCell ref="([^"]+)"/>', repl_merge, sheet_xml)


# ---------------------------------------------------------------------------
# 8. 시트 하나에 raw 파일 하나 반영
# ---------------------------------------------------------------------------

def apply_raw_to_sheet(zip_entries, master_path, sheet_name, raw_path, mapping, style_ids):
    state = read_target_sheet_state(master_path, sheet_name)
    last_data_row = state["last_data_row"]
    running = state["running"]
    include_ol = (last_data_row == 2)

    records, blocks_report, period = parse_raw_file(raw_path, include_ol)
    validation = validate_blocks(blocks_report, include_ol)

    if not include_ol and period and state["last_date_str"] and period[0] <= state["last_date_str"]:
        raise RuntimeError(
            f"[{sheet_name}] raw 파일의 기간({period[0]} ~ {period[1]})이 기존 마지막 거래일자"
            f"({state['last_date_str']})와 겹치거나 그보다 앞섭니다. 같은 기간을 두 번 넣으려는 "
            f"것일 수 있습니다."
        )

    if not records:
        return {
            "sheet": sheet_name, "added": 0, "validation": validation, "period": period,
            "include_ol": include_ol, "warnings": ["새로 추가할 거래가 없습니다."],
        }

    sheet_xml_path = find_sheet_xml_path(zip_entries, sheet_name)
    sheet_xml = zip_entries[sheet_xml_path].decode("utf-8")

    new_row_start = last_data_row + 1
    new_rows_xml, warnings = build_new_rows(new_row_start, records, style_ids, mapping, running, sheet_name)
    n_added = len(records)

    end_row_of_sheet = max(int(x) for x in re.findall(r'<row r="(\d+)"', sheet_xml))
    last_row_match = re.search(rf'<row r="{last_data_row}"[^>]*>.*?</row>', sheet_xml, re.S)
    insert_at = last_row_match.end()

    if end_row_of_sheet > last_data_row:
        m_end = re.search(r"</sheetData>", sheet_xml)
        trailing_old_text = sheet_xml[insert_at:m_end.start()]
        trailing_new_text = shift_trailing_block(trailing_old_text, last_data_row + 1, end_row_of_sheet, n_added)
        sheet_xml = sheet_xml[:insert_at] + new_rows_xml + trailing_new_text + sheet_xml[m_end.start():]
    else:
        sheet_xml = sheet_xml[:insert_at] + new_rows_xml + sheet_xml[insert_at:]

    new_last_row = end_row_of_sheet + n_added

    def repl_dim(m):
        return f'<dimension ref="{m.group(1)}{new_last_row}"/>'
    sheet_xml = re.sub(r'<dimension ref="([A-Z0-9\$:]+?)\d+"/>', repl_dim, sheet_xml)

    def repl_af(m):
        return f'<autoFilter ref="{m.group(1)}{last_data_row + n_added}"'
    sheet_xml = re.sub(r'<autoFilter ref="([A-Z0-9\$:]+?)\d+"', repl_af, sheet_xml)
    sheet_xml = re.sub(
        r'<autoFilter ref="([^"]+)">.*?</autoFilter>', r'<autoFilter ref="\1"/>', sheet_xml, flags=re.S
    )
    sheet_xml = re.sub(r'(<row r="\d+"[^>]*?) hidden="1"', r"\1", sheet_xml)

    if end_row_of_sheet > last_data_row:
        sheet_xml = shift_merge_cells(sheet_xml, last_data_row + 1, end_row_of_sheet, n_added)

    zip_entries[sheet_xml_path] = sheet_xml.encode("utf-8")

    wb_xml = zip_entries["xl/workbook.xml"].decode("utf-8")
    pattern = re.compile(rf'({re.escape(sheet_name)}!\$[A-Z]+\$2:\$[A-Z]+\$)(\d+)')
    wb_xml = pattern.sub(lambda m: f"{m.group(1)}{last_data_row + n_added}", wb_xml)
    zip_entries["xl/workbook.xml"] = wb_xml.encode("utf-8")

    return {
        "sheet": sheet_name, "added": n_added, "validation": validation, "period": period,
        "include_ol": include_ol, "warnings": warnings,
        "new_row_range": (new_row_start, new_row_start + n_added - 1),
    }


# ---------------------------------------------------------------------------
# 9. 피벗테이블 완전 제거
# ---------------------------------------------------------------------------

def remove_pivot_tables(zip_entries, infolist):
    pivot_paths = {
        n for n in zip_entries
        if n.startswith("xl/pivotTables/") or n.startswith("xl/pivotCache/")
    }
    if not pivot_paths:
        return infolist

    for p in pivot_paths:
        del zip_entries[p]
    infolist = [i for i in infolist if i.filename not in pivot_paths]

    wb_xml = zip_entries["xl/workbook.xml"].decode("utf-8")
    wb_xml = re.sub(r"<pivotCaches>.*?</pivotCaches>", "", wb_xml, flags=re.S)
    zip_entries["xl/workbook.xml"] = wb_xml.encode("utf-8")

    rels_xml = zip_entries["xl/_rels/workbook.xml.rels"].decode("utf-8")
    rels_xml = re.sub(r'<Relationship[^>]*[Pp]ivot[Cc]ache[Dd]efinition[^>]*/>', "", rels_xml)
    zip_entries["xl/_rels/workbook.xml.rels"] = rels_xml.encode("utf-8")

    ct_xml = zip_entries["[Content_Types].xml"].decode("utf-8")
    ct_xml = re.sub(r'<Override[^>]*[Pp]ivot[^>]*/>', "", ct_xml)
    zip_entries["[Content_Types].xml"] = ct_xml.encode("utf-8")

    # 시트별 rels에서 pivotTable 관계 제거 (comments/vmlDrawing 등 다른 관계는 유지)
    for n in list(zip_entries):
        if n.startswith("xl/worksheets/_rels/") and n.endswith(".rels"):
            rels = zip_entries[n].decode("utf-8")
            if "pivotTable" not in rels:
                continue
            new_rels = re.sub(r'<Relationship[^>]*[Pp]ivotTable[^>]*/>', "", rels)
            zip_entries[n] = new_rels.encode("utf-8")

    # 각 시트 1행에 남아있던 피벗 헤더 잔여 텍스트("행 레이블" 등) 셀 제거
    sst_xml = zip_entries.get("xl/sharedStrings.xml", b"").decode("utf-8")
    target_texts = {"행 레이블", "합계 : 차변금액", "합계 : 대변금액"}
    sis = re.findall(r"<si>.*?</si>", sst_xml, re.S)
    target_idx = set()
    for i, si in enumerate(sis):
        text = "".join(re.findall(r"<t[^>]*>([^<]*)</t>", si))
        if text in target_texts:
            target_idx.add(str(i))

    if target_idx:
        idx_pattern = "|".join(sorted(target_idx))
        cell_re = re.compile(rf'<c r="[A-Z]+1" s="\d+" t="s"><v>({idx_pattern})</v></c>')
        for n in list(zip_entries):
            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
                xml = zip_entries[n].decode("utf-8")
                new_xml = cell_re.sub("", xml)
                if new_xml != xml:
                    zip_entries[n] = new_xml.encode("utf-8")

    return infolist


def drop_calc_chain(zip_entries, infolist):
    calc_path = "xl/calcChain.xml"
    if calc_path not in zip_entries:
        return infolist
    del zip_entries[calc_path]
    infolist = [i for i in infolist if i.filename != calc_path]

    ct_path = "[Content_Types].xml"
    ct = zip_entries[ct_path].decode("utf-8")
    ct = re.sub(r"<Override[^>]*calcChain[^>]*/>", "", ct)
    zip_entries[ct_path] = ct.encode("utf-8")

    rels_path = "xl/_rels/workbook.xml.rels"
    rels = zip_entries[rels_path].decode("utf-8")
    rels = re.sub(r"<Relationship[^>]*calcChain[^>]*/>", "", rels)
    zip_entries[rels_path] = rels.encode("utf-8")

    return infolist


# ---------------------------------------------------------------------------
# 10. 진입점
# ---------------------------------------------------------------------------

def find_files(dir_path, label, required=True):
    files = [f for f in glob.glob(os.path.join(dir_path, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if required and len(files) == 0:
        raise RuntimeError(f"{label} 폴더({dir_path})에 xlsx 파일이 없습니다.")
    return files


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    master_files = find_files(MASTER_DIR, "master")
    if len(master_files) > 1:
        raise RuntimeError(f"master 폴더에는 파일이 1개만 있어야 합니다. 현재: {[os.path.basename(f) for f in master_files]}")
    master_path = master_files[0]
    mapping = load_mapping(master_path)
    print(f"[master] {os.path.basename(master_path)} (계정과목표 {len(mapping)}개)")

    raw_files = find_files(RAW_DIR, "raw")
    print(f"[raw 파일 {len(raw_files)}개]")
    for f in raw_files:
        print(f"  - {os.path.basename(f)}")
    print()

    zf = zipfile.ZipFile(master_path, "r")
    zip_entries = {name: zf.read(name) for name in zf.namelist()}
    infolist = zf.infolist()
    zf.close()

    style_ids = ensure_data_styles(zip_entries)

    reports = []
    for raw_path in raw_files:
        sheet_name = os.path.splitext(os.path.basename(raw_path))[0]
        print(f"=== {sheet_name} 처리 중 (raw: {os.path.basename(raw_path)}) ===")
        report = apply_raw_to_sheet(zip_entries, master_path, sheet_name, raw_path, mapping, style_ids)
        reports.append(report)

    infolist = remove_pivot_tables(zip_entries, infolist)
    infolist = drop_calc_chain(zip_entries, infolist)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    master_stem = os.path.splitext(os.path.basename(master_path))[0]
    output_path = os.path.join(OUTPUT_DIR, f"{master_stem}_filled_{timestamp}.xlsx")

    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as out_zf:
        for info in infolist:
            out_zf.writestr(info.filename, zip_entries[info.filename])

    print()
    print("=" * 60)
    print("결과 요약")
    print("=" * 60)
    for r in reports:
        print(f"\n[{r['sheet']} 시트]")
        if r["period"]:
            print(f"  raw 파일 기간: {r['period'][0]} ~ {r['period'][1]}")
        print(f"  이월잔액 포함 여부: {'포함 (시트가 비어있었음)' if r['include_ol'] else '제외 (기존 데이터 있음)'}")
        print(f"  추가된 거래: {r['added']}건")
        if "new_row_range" in r:
            print(f"  추가된 행 위치: {r['new_row_range'][0]}행 ~ {r['new_row_range'][1]}행")
        ok_count = sum(1 for v in r["validation"] if v["ok"])
        fail = [v for v in r["validation"] if not v["ok"]]
        print(f"  계정블록 검증: {ok_count}/{len(r['validation'])} 통과")
        for v in fail:
            print(f"    [불일치] 계정 {v['계정코드']}({v['계정명']}): "
                  f"담은 차변 {v['kept_debit']} vs 기대값 {v['expected_debit']}, "
                  f"담은 대변 {v['kept_credit']} vs 기대값 {v['expected_credit']}")
        for w in r["warnings"]:
            print(f"    [경고] {w}")

    print()
    print("피벗테이블 및 관련 캐시를 모두 제거했습니다.")
    print(f"결과 파일: {output_path}")
    print("원본 master 파일은 수정되지 않았습니다.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[오류] {e}")
        sys.exit(1)
