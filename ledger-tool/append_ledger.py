#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ERP 원장(raw) 엑셀을 회사별 누적 워크북의 해당 시트에 이어붙이는 도구.

폴더 구조 (이 파일과 같은 위치):
  input_raw/  - ERP에서 받은 raw 원장 xlsx 파일들을 넣는 곳.
                파일 이름 = 채워 넣을 시트 이름 (예: HT.xlsx -> "HT" 시트)
  master/     - 누적 관리 중인 워크북 파일 1개 (원본, 이 스크립트는 이 파일을
                직접 수정하지 않고 output/에 새 파일로 결과를 저장합니다)
  output/     - 결과 파일이 저장되는 곳

사용법:
  python3 append_ledger.py
(run_mac.command / run_windows.bat 더블클릭으로도 실행 가능)
"""

import glob
import os
import re
import sys
import zipfile
import datetime

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "input_raw")
MASTER_DIR = os.path.join(BASE_DIR, "master")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

MAPPING_SHEET_NAME = "99.이카운트 계정과목표_HF"
FOOTER_MARKER = "<차이내역>"

# raw 원장 파일의 컬럼 순서 (ERP 다운로드 그대로)
RAW_COLS = [
    "일자-No", "계정코드", "계정명", "거래처코드", "거래처명", "적요",
    "차변금액", "대변금액", "잔액", "부서명", "최초작성일자", "최초작성자",
    "최종수정일자", "최종수정자", "채권채무번호", "만기일자", "은행", "계좌번호",
    "예금주명", "상대계정코드", "상대계정명", "상대거래처코드", "상대거래처명",
]

# master 시트의 컬럼 순서 (A~AC, 29개). raw_index=None 인 항목은 raw에 없는 파생 컬럼.
# kind: 'text' | 'number' | 'formula_y' | 'formula_lookup' | 'lookup_static'
MASTER_COLS = [
    ("일자-No", "raw:일자-No", "text"),
    ("날짜", "derived:날짜", "text"),
    ("계정코드", "raw:계정코드", "number"),
    ("계정명", "raw:계정명", "text"),
    ("거래처코드", "raw:거래처코드", "text"),
    ("거래처명", "raw:거래처명", "text"),
    ("적요", "raw:적요", "text"),
    ("차변금액", "raw:차변금액", "number"),
    ("대변금액", "raw:대변금액", "number"),
    ("잔액", "raw:잔액", "number"),
    ("부서명", "raw:부서명", "text"),
    ("최초작성일자", "raw:최초작성일자", "text"),
    ("최초작성자", "raw:최초작성자", "text"),
    ("최종수정일자", "raw:최종수정일자", "text"),
    ("최종수정자", "raw:최종수정자", "text"),
    ("채권채무번호", "raw:채권채무번호", "text"),
    ("만기일자", "raw:만기일자", "text"),
    ("은행", "raw:은행", "text"),
    ("계좌번호", "raw:계좌번호", "text"),
    ("예금주명", "raw:예금주명", "text"),
    ("상대계정코드", "raw:상대계정코드", "text"),
    ("상대계정명", "raw:상대계정명", "text"),
    ("상대거래처코드", "raw:상대거래처코드", "text"),
    ("상대거래처명", "raw:상대거래처명", "text"),
    ("잔액2", "derived:잔액2", "number"),
    ("구분1", "derived:구분1", "formula_lookup_E"),
    ("구분2", "derived:구분2", "formula_lookup_D"),
    ("상대구분", "derived:상대구분", "lookup_static"),
    ("비고", None, "text"),
]
COL_LETTERS = [
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
    "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC",
]

DATE_NO_RE = re.compile(r"^\d{4}/\d{2}/\d{2}\s*-")
PERIOD_RE = re.compile(r"(\d{4}/\d{2}/\d{2})\s*~\s*(\d{4}/\d{2}/\d{2})")


def esc(s):
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def norm_code(v):
    """계정코드/상대계정코드를 매핑표와 비교 가능한 문자열로 정규화."""
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        return str(int(float(s)))
    except ValueError:
        return s


# ---------------------------------------------------------------------------
# 1. 매핑표 로딩 (99.이카운트 계정과목표_HF) - openpyxl read_only (가볍고 빠름)
# ---------------------------------------------------------------------------

def load_mapping(master_path):
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    if MAPPING_SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(f"매핑 시트 '{MAPPING_SHEET_NAME}'을(를) master 파일에서 찾지 못했습니다.")
    ws = wb[MAPPING_SHEET_NAME]
    mapping = {}
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        code = norm_code(row[0])
        if code is None:
            continue
        gubun2 = row[3]  # D열: 계정과목 (구분2)
        gubun1 = row[4]  # E열: 특관자 추출계정 (구분1)
        mapping[code] = {"구분1": gubun1, "구분2": gubun2}
    wb.close()
    return mapping


# ---------------------------------------------------------------------------
# 2. raw ERP 원장 파일 파싱
# ---------------------------------------------------------------------------

def parse_raw_file(raw_path):
    """반환: (transactions, blocks_report)
    transactions: dict 리스트 (RAW_COLS 키)
    blocks_report: 계정블록별 검증용 dict 리스트
    """
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    ws = wb.worksheets[0]

    transactions = []
    blocks_report = []
    header_idx = None
    block = None
    title_code, title_name = None, None
    title_re = re.compile(r"(\d+)\(([^)]+)\)\s*$")

    def flush_block():
        if block is not None:
            blocks_report.append(block)

    for row in ws.iter_rows(values_only=True):
        first_cell = row[0]
        if isinstance(first_cell, str) and first_cell.startswith("회사명"):
            m = title_re.search(first_cell)
            if m:
                title_code, title_name = m.group(1), m.group(2)
            continue

        if isinstance(first_cell, str) and first_cell.strip() == "일자-No":
            flush_block()
            header_idx = {name: i for i, name in enumerate(row) if name}
            block = {
                "계정코드": title_code, "계정명": title_name,
                "kept_debit": 0, "kept_credit": 0,
                "ol_debit": 0, "ol_credit": 0,
                "stated_debit": None, "stated_credit": None,
                "has_summary_row": False,
            }
            continue

        if header_idx is None:
            continue  # 아직 첫 헤더를 못 만남 (타이틀 행 등)

        no_val = first_cell
        row_jeokyo = row[header_idx["적요"]] if "적요" in header_idx else None
        is_opening_balance = isinstance(row_jeokyo, str) and row_jeokyo.strip() == "이월잔액"

        if is_opening_balance:
            # 이월잔액(기초 이월) 행: 일자-No가 비어있는 경우가 많아 날짜 패턴과 무관하게 먼저 판별한다.
            # 12/31 기말잔액을 새로 시작하는 경우가 아니라면 이미 기존 누적 데이터에 반영되어 있으므로 제외한다.
            rec = {name: row[header_idx[name]] if name in header_idx else None for name in RAW_COLS}
            block["ol_debit"] += rec["차변금액"] or 0
            block["ol_credit"] += rec["대변금액"] or 0
        elif isinstance(no_val, str) and DATE_NO_RE.match(no_val):
            rec = {name: row[header_idx[name]] if name in header_idx else None for name in RAW_COLS}
            debit = rec["차변금액"] or 0
            credit = rec["대변금액"] or 0
            transactions.append(rec)
            block["kept_debit"] += debit
            block["kept_credit"] += credit
        elif isinstance(no_val, str) and no_val.strip() == "합계":
            debit_idx = header_idx.get("차변금액")
            credit_idx = header_idx.get("대변금액")
            block["stated_debit"] = row[debit_idx] if debit_idx is not None else None
            block["stated_credit"] = row[credit_idx] if credit_idx is not None else None
            block["has_summary_row"] = True
        # 그 외 (타이틀행/월계행/타임스탬프행/빈행)는 무시

    flush_block()

    # 파일 상단 타이틀에서 기간(period) 추출
    period = None
    ws0 = wb.worksheets[0]
    for row in ws0.iter_rows(min_row=1, max_row=1, values_only=True):
        if row and row[0]:
            m = PERIOD_RE.search(str(row[0]))
            if m:
                period = (m.group(1), m.group(2))
        break

    wb.close()
    return transactions, blocks_report, period


def validate_blocks(blocks_report):
    """블록별로 (합계 - 이월잔액) == 실제로 담은 거래 합계 인지 검증.
    해당 기간에 실거래가 하나도 없는 계정은 ERP가 '합계' 행 자체를 찍지 않으므로
    그 경우는 kept 합계가 0인지만 확인한다 (합계 행이 있지만 순증감이 0이라
    차변/대변 값이 비어있는 경우와는 구분해야 한다).
    """
    results = []
    for b in blocks_report:
        if not b["has_summary_row"]:
            expected_d = 0
            expected_c = 0
        else:
            expected_d = (b["stated_debit"] or 0) - b["ol_debit"]
            expected_c = (b["stated_credit"] or 0) - b["ol_credit"]
        ok = (expected_d == b["kept_debit"]) and (expected_c == b["kept_credit"])
        results.append({
            "계정코드": b["계정코드"], "계정명": b["계정명"], "ok": ok,
            "kept_debit": b["kept_debit"], "kept_credit": b["kept_credit"],
            "expected_debit": expected_d, "expected_credit": expected_c,
        })
    return results


# ---------------------------------------------------------------------------
# 3. master 시트의 기존 상태 읽기 (openpyxl read_only - 대상 시트만 접근하므로 가벼움)
# ---------------------------------------------------------------------------

def read_target_sheet_state(master_path, sheet_name):
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"master 파일에 '{sheet_name}' 시트가 없습니다. 사용 가능한 시트: {wb.sheetnames}")
    ws = wb[sheet_name]

    last_data_row = None
    marker_row = None
    running = {}  # 계정명별 (누적차변, 누적대변) - Y열(잔액2) 수식 재현용
    last_date_str = None

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        a = row[0]
        if isinstance(a, str) and a.strip() == FOOTER_MARKER:
            marker_row = i
            break
        if isinstance(a, str) and DATE_NO_RE.match(a):
            last_data_row = i
            last_date_str = row[1]
            acct_name, debit, credit = row[3], row[7], row[8]
            cur = running.get(acct_name, [0, 0])
            cur[0] += (debit or 0)
            cur[1] += (credit or 0)
            running[acct_name] = cur

    wb.close()
    if last_data_row is None:
        raise RuntimeError(f"'{sheet_name}' 시트에서 기존 거래 데이터를 찾지 못했습니다.")
    return {
        "last_data_row": last_data_row,
        "marker_row": marker_row,
        "running": running,
        "last_date_str": last_date_str,
    }


# ---------------------------------------------------------------------------
# 4. master workbook 안의 sheet XML 경로 찾기
# ---------------------------------------------------------------------------

def find_sheet_xml_path(zf, sheet_name):
    wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
    m = re.search(rf'<sheet name="{re.escape(sheet_name)}"[^>]*r:id="(rId\d+)"', wb_xml)
    if not m:
        raise RuntimeError(f"workbook.xml에서 '{sheet_name}' 시트를 찾지 못했습니다.")
    rid = m.group(1)
    rels_path = "xl/_rels/workbook.xml.rels"
    rels_xml = zf.read(rels_path).decode("utf-8")
    m2 = re.search(rf'<Relationship Id="{rid}"[^>]*Target="([^"]+)"', rels_xml)
    if not m2:
        raise RuntimeError(f"workbook.xml.rels에서 {rid}를 찾지 못했습니다.")
    target = m2.group(1)
    return "xl/" + target.lstrip("/")


# ---------------------------------------------------------------------------
# 5. 스타일 템플릿 추출 (기존 시트 원본 xml에서 컬럼별 style id 가져오기)
# ---------------------------------------------------------------------------

def extract_row_inner(sheet_xml, row_num):
    m = re.search(rf'<row r="{row_num}"[^>]*>(.*?)</row>', sheet_xml, re.S)
    return m.group(1) if m else None


def cell_style_map(row_inner):
    styles = {}
    for cm in re.finditer(r'<c r="([A-Z]+)\d+"([^>]*?)(?:/>|>)', row_inner):
        col, attrs = cm.group(1), cm.group(2)
        sm = re.search(r's="(\d+)"', attrs)
        styles[col] = sm.group(1) if sm else None
    return styles


def find_style_template(sheet_xml, last_data_row):
    """Z열(구분1)에 수식이 살아있는 가장 가까운 행을 서식 템플릿으로 사용."""
    for r in range(last_data_row, 2, -1):
        inner = extract_row_inner(sheet_xml, r)
        if inner and re.search(r'<c r="Z\d+"[^>]*>\s*<f>', inner):
            styles = cell_style_map(inner)
            # AB(상대구분)는 별도로 값이 있는 행에서 스타일을 찾아본다
            if "AB" not in styles:
                for r2 in range(last_data_row, 2, -1):
                    inner2 = extract_row_inner(sheet_xml, r2)
                    if inner2:
                        s2 = cell_style_map(inner2)
                        if "AB" in s2:
                            styles["AB"] = s2["AB"]
                            break
            return r, styles
    raise RuntimeError("서식 템플릿으로 쓸 행을 찾지 못했습니다 (Z열에 수식이 있는 행이 없음).")


# ---------------------------------------------------------------------------
# 6. 새 row XML 생성
# ---------------------------------------------------------------------------

def build_cell(col_letter, row_num, style, kind, value, formula=None, cached=None):
    ref = f"{col_letter}{row_num}"
    s_attr = f' s="{style}"' if style else ""
    if kind == "text":
        if value is None or value == "":
            return f'<c r="{ref}"{s_attr}/>'
        return f'<c r="{ref}"{s_attr} t="inlineStr"><is><t xml:space="preserve">{esc(value)}</t></is></c>'
    if kind == "number":
        if value is None or value == "":
            return f'<c r="{ref}"{s_attr}/>'
        return f'<c r="{ref}"{s_attr}><v>{value}</v></c>'
    if kind == "formula_num":
        return f'<c r="{ref}"{s_attr}><f>{esc(formula)}</f><v>{cached}</v></c>'
    if kind == "formula_str":
        return f'<c r="{ref}"{s_attr} t="str"><f>{esc(formula)}</f><v>{esc(cached)}</v></c>'
    raise ValueError(kind)


def build_new_rows(new_row_start, records, styles, mapping, running, sheet_name):
    """records: raw에서 뽑은 거래 dict 리스트. 반환: xml 문자열, 매핑 실패 경고 리스트"""
    xml_parts = []
    warnings = []
    for offset, rec in enumerate(records):
        r = new_row_start + offset
        raw_no = rec["일자-No"]
        date_str = raw_no.split()[0] if raw_no else None
        acct_code_norm = norm_code(rec["계정코드"])
        acct_code_num = int(acct_code_norm) if acct_code_norm is not None else None
        acct_name = rec["계정명"]
        debit = rec["차변금액"] or 0
        credit = rec["대변금액"] or 0

        # Y열 (잔액2): 계정명별 누적 |차변-대변| - 기존 running 딕셔너리를 이어서 갱신
        cur = running.get(acct_name, [0, 0])
        cur[0] += debit
        cur[1] += credit
        running[acct_name] = cur
        y_value = abs(cur[0] - cur[1])
        y_formula = f"ABS(SUMIFS(H$3:H{r}, D$3:D{r}, D{r}) - SUMIFS(I$3:I{r}, D$3:D{r}, D{r}))"

        # Z/AA열 (구분1/구분2): 계정과목표 매핑 INDEX/MATCH 수식 재현
        map_entry = mapping.get(acct_code_norm)
        if map_entry is None:
            warnings.append(f"[{sheet_name}] 계정코드 {rec['계정코드']} 이(가) 매핑표에 없습니다 (행 {r}, Z/AA 공백 처리).")
            gubun1, gubun2 = "", ""
        else:
            gubun1 = map_entry["구분1"] if map_entry["구분1"] is not None else ""
            gubun2 = map_entry["구분2"] if map_entry["구분2"] is not None else ""
        z_formula = f"INDEX('{MAPPING_SHEET_NAME}'!E:E,MATCH({sheet_name}!C{r},'{MAPPING_SHEET_NAME}'!A:A,0))"
        aa_formula = f"INDEX('{MAPPING_SHEET_NAME}'!D:D,MATCH({sheet_name}!C{r},'{MAPPING_SHEET_NAME}'!A:A,0))"

        # AB열 (상대구분): 수기로 확인하는 항목이므로 자동으로 채우지 않고 공란으로 둔다.

        values_by_col = {
            "A": ("text", raw_no),
            "B": ("text", date_str),
            "C": ("number", acct_code_num),
            "D": ("text", acct_name),
            "E": ("text", rec["거래처코드"]),
            "F": ("text", rec["거래처명"]),
            "G": ("text", rec["적요"]),
            "H": ("number", rec["차변금액"]),
            "I": ("number", rec["대변금액"]),
            "J": ("number", rec["잔액"]),
            "K": ("text", rec["부서명"]),
            "L": ("text", rec["최초작성일자"]),
            "M": ("text", rec["최초작성자"]),
            "N": ("text", rec["최종수정일자"]),
            "O": ("text", rec["최종수정자"]),
            "P": ("text", rec["채권채무번호"]),
            "Q": ("text", rec["만기일자"]),
            "R": ("text", rec["은행"]),
            "S": ("text", rec["계좌번호"]),
            "T": ("text", rec["예금주명"]),
            "U": ("text", rec["상대계정코드"]),
            "V": ("text", rec["상대계정명"]),
            "W": ("text", rec["상대거래처코드"]),
            "X": ("text", rec["상대거래처명"]),
        }

        row_cells = []
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
                    "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X"]:
            kind, val = values_by_col[col]
            row_cells.append(build_cell(col, r, styles.get(col), kind, val))
        row_cells.append(build_cell("Y", r, styles.get("Y"), "formula_num", None, y_formula, y_value))
        row_cells.append(build_cell("Z", r, styles.get("Z"), "formula_str", None, z_formula, gubun1))
        row_cells.append(build_cell("AA", r, styles.get("AA"), "formula_str", None, aa_formula, gubun2))
        row_cells.append(build_cell("AB", r, styles.get("AB"), "text", None))

        xml_parts.append(f'<row r="{r}" spans="1:28" x14ac:dyDescent="0.3">' + "".join(row_cells) + "</row>")

    return "".join(xml_parts), warnings


# ---------------------------------------------------------------------------
# 7. footer 블록 이동 (마커 행부터 시트 끝까지, 행 번호를 n만큼 밀기)
# ---------------------------------------------------------------------------

def shift_footer_block(block_text, old_start, old_end, n):
    def repl_row(m):
        return f'<row r="{int(m.group(1)) + n}"'
    block_text = re.sub(r'<row r="(\d+)"', repl_row, block_text)

    nums = "|".join(str(x) for x in range(old_start, old_end + 1))
    pattern = re.compile(r"([A-Z]{1,3})(" + nums + r")\b")

    def repl_cell(m):
        return f"{m.group(1)}{int(m.group(2)) + n}"
    block_text = pattern.sub(repl_cell, block_text)
    return block_text


def shift_merge_cells(sheet_xml, old_start, old_end, n):
    nums = "|".join(str(x) for x in range(old_start, old_end + 1))
    pattern = re.compile(r"([A-Z]{1,3})(" + nums + r")\b")

    def repl(m):
        return f"{m.group(1)}{int(m.group(2)) + n}"

    def repl_merge(m):
        return "<mergeCell ref=\"" + pattern.sub(repl, m.group(1)) + "\"/>"

    return re.sub(r'<mergeCell ref="([^"]+)"/>', repl_merge, sheet_xml)


# ---------------------------------------------------------------------------
# 8. 메인 처리: 하나의 raw 파일을 master의 해당 시트에 반영
# ---------------------------------------------------------------------------

def apply_raw_to_sheet(zip_entries, master_path, sheet_name, raw_path, mapping):
    records, blocks_report, period = parse_raw_file(raw_path)
    validation = validate_blocks(blocks_report)

    state = read_target_sheet_state(master_path, sheet_name)
    last_data_row = state["last_data_row"]
    marker_row = state["marker_row"]
    running = state["running"]

    if not records:
        return {
            "sheet": sheet_name, "added": 0, "validation": validation,
            "period": period, "warnings": ["새로 추가할 실제 거래가 없습니다 (전부 이월잔액이거나 빈 파일)."],
            "last_date_before": state["last_date_str"],
        }

    sheet_xml_path = find_sheet_xml_path(zipfile.ZipFile(master_path), sheet_name)
    sheet_xml = zip_entries[sheet_xml_path].decode("utf-8")

    template_row, styles = find_style_template(sheet_xml, last_data_row)

    new_row_start = last_data_row + 1
    new_rows_xml, warnings = build_new_rows(new_row_start, records, styles, mapping, running, sheet_name)
    n_added = len(records)

    if marker_row:
        # footer 블록 원문 추출 (marker_row 부터 </sheetData> 직전까지)
        m_start = re.search(rf'<row r="{marker_row}"', sheet_xml)
        m_end = re.search(r"</sheetData>", sheet_xml)
        footer_old_text = sheet_xml[m_start.start():m_end.start()]
        old_end_row = max(int(x) for x in re.findall(r'<row r="(\d+)"', footer_old_text))

        footer_new_text = shift_footer_block(footer_old_text, marker_row, old_end_row, n_added)

        sheet_xml = (
            sheet_xml[:m_start.start()]
            + new_rows_xml
            + footer_new_text
            + sheet_xml[m_end.start():]
        )
        new_last_row = old_end_row + n_added
    else:
        # 각주 블록이 없는 시트: 마지막 데이터 행 뒤에 그냥 이어붙임
        insert_at = re.search(rf'<row r="{last_data_row}"[^>]*>.*?</row>', sheet_xml, re.S).end()
        sheet_xml = sheet_xml[:insert_at] + new_rows_xml + sheet_xml[insert_at:]
        new_last_row = last_data_row + n_added

    # dimension 갱신
    def repl_dim(m):
        return f'<dimension ref="{m.group(1)}{new_last_row}"/>'
    sheet_xml = re.sub(r'<dimension ref="([A-Z0-9\$:]+?)\d+"/>', repl_dim, sheet_xml)

    # autoFilter 갱신 (마지막 데이터 행까지 범위 확장)
    def repl_af(m):
        return f'<autoFilter ref="{m.group(1)}{last_data_row + n_added}"'
    sheet_xml = re.sub(r'<autoFilter ref="([A-Z0-9\$:]+?)\d+"', repl_af, sheet_xml)

    # mergeCells (footer 안에 있던 것들) 행 번호 이동
    if marker_row:
        sheet_xml = shift_merge_cells(sheet_xml, marker_row, old_end_row, n_added)

    zip_entries[sheet_xml_path] = sheet_xml.encode("utf-8")

    # workbook.xml 의 _xlnm._FilterDatabase 정의된 이름 범위도 갱신
    wb_xml = zip_entries["xl/workbook.xml"].decode("utf-8")
    pattern = re.compile(
        rf'({re.escape(sheet_name)}!\$[A-Z]+\$2:\$[A-Z]+\$)(\d+)'
    )
    def repl_defname(m):
        return f"{m.group(1)}{last_data_row + n_added}"
    wb_xml = pattern.sub(repl_defname, wb_xml)
    zip_entries["xl/workbook.xml"] = wb_xml.encode("utf-8")

    return {
        "sheet": sheet_name, "added": n_added, "validation": validation,
        "period": period, "warnings": warnings,
        "last_date_before": state["last_date_str"],
        "new_row_range": (new_row_start, new_row_start + n_added - 1),
    }


# ---------------------------------------------------------------------------
# 9. 진입점
# ---------------------------------------------------------------------------

def drop_calc_chain(zip_entries, infolist):
    """calcChain.xml은 엑셀이 수식 계산 순서를 캐싱해둔 파일이라, 새 수식 셀을
    추가하면 그 내용과 실제 수식 셀 목록이 어긋나 엑셀이 파일을 열 때 '내용에
    문제가 있다'며 복구 창을 띄운다. 이 파일은 없어도 엑셀이 열 때 자동으로
    다시 만들어주므로, 안전하게 빼고 관련 참조도 같이 정리한다.
    """
    calc_path = "xl/calcChain.xml"
    if calc_path not in zip_entries:
        return infolist

    del zip_entries[calc_path]
    infolist = [i for i in infolist if i.filename != calc_path]

    ct_path = "[Content_Types].xml"
    ct = zip_entries[ct_path].decode("utf-8")
    ct = re.sub(r'<Override[^>]*calcChain[^>]*/>', "", ct)
    zip_entries[ct_path] = ct.encode("utf-8")

    rels_path = "xl/_rels/workbook.xml.rels"
    rels = zip_entries[rels_path].decode("utf-8")
    rels = re.sub(r'<Relationship[^>]*calcChain[^>]*/>', "", rels)
    zip_entries[rels_path] = rels.encode("utf-8")

    return infolist


def find_single_xlsx(dir_path, label):
    files = [f for f in glob.glob(os.path.join(dir_path, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if len(files) == 0:
        raise RuntimeError(f"{label} 폴더({dir_path})에 xlsx 파일이 없습니다.")
    if len(files) > 1 and label == "master":
        raise RuntimeError(f"{label} 폴더에는 파일이 1개만 있어야 합니다. 현재: {[os.path.basename(f) for f in files]}")
    return files if label != "master" else files[0]


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    master_path = find_single_xlsx(MASTER_DIR, "master")
    raw_files = find_single_xlsx(RAW_DIR, "raw")

    print(f"[master] {os.path.basename(master_path)}")
    print(f"[raw 파일 {len(raw_files)}개]")
    for f in raw_files:
        print(f"  - {os.path.basename(f)}")
    print()

    mapping = load_mapping(master_path)

    zf = zipfile.ZipFile(master_path, "r")
    zip_entries = {name: zf.read(name) for name in zf.namelist()}
    infolist = zf.infolist()
    zf.close()

    reports = []
    for raw_path in raw_files:
        sheet_name = os.path.splitext(os.path.basename(raw_path))[0]
        print(f"=== {sheet_name} 처리 중 (raw: {os.path.basename(raw_path)}) ===")
        report = apply_raw_to_sheet(zip_entries, master_path, sheet_name, raw_path, mapping)
        reports.append(report)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    master_stem = os.path.splitext(os.path.basename(master_path))[0]
    output_path = os.path.join(OUTPUT_DIR, f"{master_stem}_updated_{timestamp}.xlsx")

    infolist = drop_calc_chain(zip_entries, infolist)

    with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as out_zf:
        for info in infolist:
            out_zf.writestr(info.filename, zip_entries[info.filename])

    print()
    print("=" * 60)
    print("결과 요약")
    print("=" * 60)
    for r in reports:
        print(f"\n[{r['sheet']} 시트]")
        if r["period"]:
            print(f"  raw 파일 기간: {r['period'][0]} ~ {r['period'][1]}")
        print(f"  기존 마지막 거래일자: {r['last_date_before']}")
        print(f"  새로 추가된 거래: {r['added']}건")
        if "new_row_range" in r:
            print(f"  추가된 행 위치: {r['new_row_range'][0]}행 ~ {r['new_row_range'][1]}행")
        ok_count = sum(1 for v in r["validation"] if v["ok"])
        fail = [v for v in r["validation"] if not v["ok"]]
        print(f"  계정블록 검증: {ok_count}/{len(r['validation'])} 통과")
        for v in fail:
            print(f"    [불일치] 계정 {v['계정코드']}({v['계정명']}): "
                  f"담은 차변 {v['kept_debit']} vs 기대값 {v['expected_debit']}, "
                  f"담은 대변 {v['kept_credit']} vs 기대값 {v['expected_credit']}")
        for w in r["warnings"]:
            print(f"    [경고] {w}")

    print()
    print(f"결과 파일: {output_path}")
    print("원본 master 파일은 수정되지 않았습니다. 내용을 확인한 뒤 직접 교체해 주세요.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[오류] {e}")
        sys.exit(1)
