#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ERP 원장(raw) 엑셀만으로, 기존 누적 워크북(거래 데이터가 담긴 master) 없이
새 워크북을 처음부터 만드는 도구.

누적형 도구(ledger-tool)와의 차이:
  - 거래 데이터가 담긴 master 워크북이 필요 없습니다. 대신 계정과목표
    (99.이카운트 계정과목표_HF) 시트만 포함된 "양식" 파일이 필요합니다 —
    계정코드가 어떤 구분(채권/채무/자산 등)에 속하는지는 raw 원장 파일에
    없고 이 표에만 있기 때문입니다.
  - 이월잔액 행을 "제외"하지 않고 "포함"합니다 (처음 만드는 것이므로 기초잔액이 필요).

폴더 구조 (이 파일과 같은 위치):
  input_raw/  - ERP에서 받은 raw 원장 xlsx 파일들. 파일명 = 만들어질 시트 이름 (예: HT.xlsx -> "HT" 시트)
  master/     - 계정과목표(99.이카운트 계정과목표_HF) 시트만 포함된 양식 파일 1개
                (거래 데이터는 없어도 됨)
  output/     - 새로 만들어진 워크북이 저장되는 곳

사용법:
  python3 create_ledger.py
(run_mac.command / run_windows.bat 더블클릭으로도 실행 가능)
"""

import datetime
import glob
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "input_raw")
MASTER_DIR = os.path.join(BASE_DIR, "master")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

MAPPING_SHEET_NAME = "99.이카운트 계정과목표_HF"

RAW_COLS = [
    "일자-No", "계정코드", "계정명", "거래처코드", "거래처명", "적요",
    "차변금액", "대변금액", "잔액", "부서명", "최초작성일자", "최초작성자",
    "최종수정일자", "최종수정자", "채권채무번호", "만기일자", "은행", "계좌번호",
    "예금주명", "상대계정코드", "상대계정명", "상대거래처코드", "상대거래처명",
]

MASTER_HEADER = [
    "일자-No", "날짜", "계정코드", "계정명", "거래처코드", "거래처명", "적요",
    "차변금액", "대변금액", "잔액", "부서명", "최초작성일자", "최초작성자",
    "최종수정일자", "최종수정자", "채권채무번호", "만기일자", "은행", "계좌번호",
    "예금주명", "상대계정코드", "상대계정명", "상대거래처코드", "상대거래처명",
    "잔액", "구분1", "구분2", "상대구분", "비고",
]

DATE_NO_RE = re.compile(r"^\d{4}/\d{2}/\d{2}\s*-")
PERIOD_RE = re.compile(r"(\d{4}/\d{2}/\d{2})\s*~\s*(\d{4}/\d{2}/\d{2})")


def norm_code(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        return str(int(float(s)))
    except ValueError:
        return s


def parse_date_str(s):
    if not s:
        return None
    try:
        y, m, d = (int(x) for x in s.split("/"))
        return datetime.date(y, m, d)
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# 매핑표 로딩
# ---------------------------------------------------------------------------

def load_mapping(master_template_path):
    wb = openpyxl.load_workbook(master_template_path, read_only=True, data_only=True)
    if MAPPING_SHEET_NAME not in wb.sheetnames:
        raise RuntimeError(
            f"master 폴더의 양식 파일에서 '{MAPPING_SHEET_NAME}' 시트를 찾지 못했습니다. "
            f"이 파일에 있는 시트: {wb.sheetnames}"
        )
    ws = wb[MAPPING_SHEET_NAME]
    mapping = {}
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        code = norm_code(row[0])
        if code is None:
            continue
        mapping[code] = {"구분1": row[4], "구분2": row[3]}
    wb.close()
    return mapping


# ---------------------------------------------------------------------------
# raw ERP 원장 파일 파싱 (이월잔액 포함 버전)
# ---------------------------------------------------------------------------

def parse_raw_file(raw_path):
    wb = openpyxl.load_workbook(raw_path, data_only=True)
    ws = wb.worksheets[0]

    transactions = []
    blocks_report = []
    header_idx = None
    block = None
    title_code, title_name = None, None
    period = None
    title_re = re.compile(r"(\d+)\(([^)]+)\)\s*$")

    def flush_block():
        if block is not None:
            blocks_report.append(block)

    for row in ws.iter_rows(values_only=True):
        first_cell = row[0]
        if isinstance(first_cell, str) and first_cell.startswith("회사명"):
            if period is None:
                pm = PERIOD_RE.search(first_cell)
                if pm:
                    period = (pm.group(1), pm.group(2))
            m = title_re.search(first_cell)
            if m:
                title_code, title_name = m.group(1), m.group(2)
            continue

        if isinstance(first_cell, str) and first_cell.strip() == "일자-No":
            flush_block()
            header_idx = {name: i for i, name in enumerate(row) if name}
            block = {
                "계정코드": title_code, "계정명": title_name,
                "kept_debit": 0, "kept_credit": 0,
                "ol_debit": 0, "ol_credit": 0,
                "stated_debit": None, "stated_credit": None,
                "has_summary_row": False,
            }
            continue

        if header_idx is None:
            continue

        no_val = first_cell
        row_jeokyo = row[header_idx["적요"]] if "적요" in header_idx else None
        is_opening_balance = isinstance(row_jeokyo, str) and row_jeokyo.strip() == "이월잔액"

        if is_opening_balance:
            rec = {name: row[header_idx[name]] if name in header_idx else None for name in RAW_COLS}
            block["ol_debit"] += rec["차변금액"] or 0
            block["ol_credit"] += rec["대변금액"] or 0
            # 새로 만드는 워크북이므로 이월잔액도 기초잔액 거래 행으로 포함한다.
            if period:
                rec["일자-No"] = f"{period[0]} -000"
            rec["계정코드"] = rec["계정코드"] or title_code
            rec["계정명"] = rec["계정명"] or title_name
            transactions.append(rec)
            block["kept_debit"] += rec["차변금액"] or 0
            block["kept_credit"] += rec["대변금액"] or 0
        elif isinstance(no_val, str) and DATE_NO_RE.match(no_val):
            rec = {name: row[header_idx[name]] if name in header_idx else None for name in RAW_COLS}
            debit = rec["차변금액"] or 0
            credit = rec["대변금액"] or 0
            transactions.append(rec)
            block["kept_debit"] += debit
            block["kept_credit"] += credit
        elif isinstance(no_val, str) and no_val.strip() == "합계":
            debit_idx = header_idx.get("차변금액")
            credit_idx = header_idx.get("대변금액")
            block["stated_debit"] = row[debit_idx] if debit_idx is not None else None
            block["stated_credit"] = row[credit_idx] if credit_idx is not None else None
            block["has_summary_row"] = True

    flush_block()
    wb.close()
    return transactions, blocks_report, period


def validate_blocks(blocks_report):
    """이월잔액을 포함해서 담았으므로:
    - '합계' 행이 있으면 그 값 그대로가 기대값 (이월잔액+실거래를 합친 값이므로).
    - '합계' 행이 없으면(그 기간에 실거래가 없었던 계정) 이월잔액만 담겼어야 한다.
    """
    results = []
    for b in blocks_report:
        if not b["has_summary_row"]:
            expected_d, expected_c = b["ol_debit"], b["ol_credit"]
        else:
            expected_d = b["stated_debit"] or 0
            expected_c = b["stated_credit"] or 0
        ok = (expected_d == b["kept_debit"]) and (expected_c == b["kept_credit"])
        results.append({
            "계정코드": b["계정코드"], "계정명": b["계정명"], "ok": ok,
            "kept_debit": b["kept_debit"], "kept_credit": b["kept_credit"],
            "expected_debit": expected_d, "expected_credit": expected_c,
        })
    return results


# ---------------------------------------------------------------------------
# 워크북 생성
# ---------------------------------------------------------------------------

FONT_NORMAL = Font(name="Arial", size=10)
FONT_HEADER = Font(name="Arial", size=10, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = "#,##0;[Red]\\(#,##0\\)"
DATE_FMT = "yyyy/mm/dd"


def build_sheet(wb, sheet_name, records, mapping, warnings):
    ws = wb.create_sheet(title=sheet_name)

    for col_idx, name in enumerate(MASTER_HEADER, start=1):
        c = ws.cell(row=2, column=col_idx, value=name)
        c.font = FONT_HEADER
        c.border = BORDER_ALL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(MASTER_HEADER))}2"

    running = {}
    for offset, rec in enumerate(records):
        r = 3 + offset
        raw_no = rec["일자-No"]
        date_val = parse_date_str(raw_no.split()[0] if raw_no else None)
        acct_code_norm = norm_code(rec["계정코드"])
        acct_code_num = int(acct_code_norm) if acct_code_norm is not None else None
        acct_name = rec["계정명"]
        debit = rec["차변금액"] or 0
        credit = rec["대변금액"] or 0

        cur = running.get(acct_name, [0, 0])
        cur[0] += debit
        cur[1] += credit
        running[acct_name] = cur
        y_value = abs(cur[0] - cur[1])

        map_entry = mapping.get(acct_code_norm)
        if map_entry is None:
            warnings.append(f"[{sheet_name}] 계정코드 {rec['계정코드']} 이(가) 매핑표에 없습니다 (행 {r}, 구분1/구분2 공란 처리).")
            gubun1, gubun2 = "", ""
        else:
            gubun1 = map_entry["구분1"] if map_entry["구분1"] is not None else ""
            gubun2 = map_entry["구분2"] if map_entry["구분2"] is not None else ""

        values = [
            raw_no, date_val, acct_code_num, acct_name, rec["거래처코드"], rec["거래처명"], rec["적요"],
            rec["차변금액"], rec["대변금액"], rec["잔액"], rec["부서명"], rec["최초작성일자"], rec["최초작성자"],
            rec["최종수정일자"], rec["최종수정자"], rec["채권채무번호"], rec["만기일자"], rec["은행"],
            rec["계좌번호"], rec["예금주명"], rec["상대계정코드"], rec["상대계정명"], rec["상대거래처코드"],
            rec["상대거래처명"], y_value, gubun1, gubun2, None, None,
        ]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_idx, value=val)
            c.font = FONT_NORMAL
            c.border = BORDER_ALL
            col_letter = get_column_letter(col_idx)
            if col_letter == "B":
                c.number_format = DATE_FMT
            elif col_letter in ("H", "I", "J", "Y"):
                c.number_format = MONEY_FMT

    for col_idx in range(1, len(MASTER_HEADER) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    return ws


def find_files(dir_path, label, required=True):
    files = [f for f in glob.glob(os.path.join(dir_path, "*.xlsx")) if not os.path.basename(f).startswith("~$")]
    if required and len(files) == 0:
        raise RuntimeError(f"{label} 폴더({dir_path})에 xlsx 파일이 없습니다.")
    return files


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    master_files = find_files(MASTER_DIR, "master")
    if len(master_files) > 1:
        raise RuntimeError(f"master 폴더에는 파일이 1개만 있어야 합니다. 현재: {[os.path.basename(f) for f in master_files]}")
    mapping = load_mapping(master_files[0])
    print(f"[master 양식] {os.path.basename(master_files[0])} ({len(mapping)}개 계정코드)")

    raw_files = find_files(RAW_DIR, "raw")
    print(f"[raw 파일 {len(raw_files)}개]")
    for f in raw_files:
        print(f"  - {os.path.basename(f)}")
    print()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 기본 빈 시트 제거

    reports = []
    for raw_path in raw_files:
        sheet_name = os.path.splitext(os.path.basename(raw_path))[0]
        print(f"=== {sheet_name} 처리 중 (raw: {os.path.basename(raw_path)}) ===")
        records, blocks_report, period = parse_raw_file(raw_path)
        validation = validate_blocks(blocks_report)
        warnings = []
        if records:
            build_sheet(wb, sheet_name, records, mapping, warnings)
        reports.append({
            "sheet": sheet_name, "added": len(records), "validation": validation,
            "period": period, "warnings": warnings,
        })

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(OUTPUT_DIR, f"신규_거래내역_{timestamp}.xlsx")
    wb.save(output_path)

    print()
    print("=" * 60)
    print("결과 요약")
    print("=" * 60)
    for r in reports:
        print(f"\n[{r['sheet']} 시트]")
        if r["period"]:
            print(f"  raw 파일 기간: {r['period'][0]} ~ {r['period'][1]}")
        print(f"  담은 거래(이월잔액 포함): {r['added']}건")
        ok_count = sum(1 for v in r["validation"] if v["ok"])
        fail = [v for v in r["validation"] if not v["ok"]]
        print(f"  계정블록 검증: {ok_count}/{len(r['validation'])} 통과")
        for v in fail:
            print(f"    [불일치] 계정 {v['계정코드']}({v['계정명']}): "
                  f"담은 차변 {v['kept_debit']} vs 기대값 {v['expected_debit']}, "
                  f"담은 대변 {v['kept_credit']} vs 기대값 {v['expected_credit']}")
        for w in r["warnings"]:
            print(f"    [경고] {w}")

    print()
    print(f"결과 파일: {output_path}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n[오류] {e}")
        sys.exit(1)
